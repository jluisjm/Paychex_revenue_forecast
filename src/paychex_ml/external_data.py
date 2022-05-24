import json, os
import requests
import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook

#from azure.storage.blob import BlobServiceClient
from src.paychex_ml.utils import load_credentials
from src.paychex_ml.utils import upload_df_parquet, upload_df_csv


def get_bls_data(seriesid, startyear='2015', endyear='2022'):
    """
    :param seriesid:
    :param startyear:
    :param endyear:
    :return:
    """

    # Names dictionary
    series_dic = {"LNS14000000": 'UnemploymentRate',
                  "CES0000000001": "NFPayrolls_sa",
                  "CEU0000000001": "NFPayrolls_nsa"}

    # Request data
    headers = {'Content-type': 'application/json'}
    data = json.dumps({"seriesid": seriesid, "startyear": startyear, "endyear": endyear})
    p = requests.post('https://api.bls.gov/publicAPI/v2/timeseries/data/', data=data, headers=headers)

    # Load data
    json_response = json.loads(p.text)
    json_data = json_response['Results']['series'][0]['data']

    # Data as dataframe
    list_df = []
    for s in json_response['Results']['series']:

        print("Processing: ", s['seriesID'])

        try:
            col_name = series_dic[s['seriesID']]
        except:
            col_name = s['seriesID']

        ser = pd.DataFrame(s['data']) \
            .astype({'value': 'float'}) \
            .drop(columns=['latest', 'footnotes', 'period']) \
            .rename(columns={'value': col_name}) \
            .set_index(['year', 'periodName'])
        list_df.append(ser)

    df = pd.concat(list_df, axis=1).reset_index()
    df['date'] = pd.to_datetime(df['year'] + df['periodName'] + '01', format="%Y%B%d").dt.strftime("%Y%m%d")
    df = df.drop(columns=['year', 'periodName'])
    #df = df[['date'] + list(filter(lambda x: x != 'date', df.columns.values))]
    return df.set_index('date')

def get_census_data(startyear='2015'):
    '''

    :param startyear:
    :return:
    '''

    print("Reading census data")
    url = 'http://api.census.gov/data/timeseries/eits/bfs?get=cell_value,time_slot_id&category_code=TOTAL&seasonally_adj&data_type_code=BA_BA&for=US&time=from+{}'\
        .format(startyear)
    p = requests.get(url)
    data = json.loads(p.text)

    df = pd.DataFrame(data)
    df = df.rename(columns=df.iloc[0]).drop(df.index[0])
    df['date'] = pd.to_datetime(df['time'] + '01', format="%Y-%m%d").dt.strftime("%Y%m%d")
    df = df.drop(['time_slot_id','category_code','data_type_code','time','us'], axis=1)\
        .set_index(['date','seasonally_adj'])\
        .unstack(1)['cell_value']\
        .rename(columns={'no': 'BusinessApplications_nsa', 'yes': 'BusinessApplications_sa'})

    return df

def get_external_data(seriesbls, startyear='2015', endyear='2022'):

    df_bls = get_bls_data(seriesbls, startyear, endyear)
    df_census = get_census_data(startyear)

    df = pd.concat([df_bls, df_census], axis=1)
    df = df.reset_index().rename(columns={'index': 'date'})

    return df

def get_fred_series(id_series):

    print("Loading series: ", id_series)
    url = 'https://api.stlouisfed.org/fred/series/observations?series_id={}&observation_start=2000-01-01&api_key={}&file_type=json' \
        .format(id_series,"2f07bf8c1db37581bdb4874f8fa68418")
    p = requests.get(url)
    json_response = json.loads(p.text)
    df = pd.DataFrame(json_response['observations']) \
        .rename(columns={'value': id_series}) \
        .drop(columns=['realtime_start', 'realtime_end'])

    df['date'] = pd.to_datetime(df['date'],format="%Y-%m-%d")

    df = df.set_index('date') \
        .replace(".", np.nan) \
        .astype(float) \
        .resample('MS').mean() \
        .interpolate('time')

    df = df[df.index>= '2010-01-01']
    return df


def get_fred_data(series_dict, write_excel=False, path='external_data.xlsx'):

    total_df = []

    if write_excel:
        book = Workbook()
        book.save(path)

    for k in series_dict:
        series_df = []
        print("Loading category: ", k)

        for s in series_dict[k]:
            series_df.append(get_fred_series(s))

        df_cat = pd.concat(series_df, axis=1)
        df_cat = df_cat.reset_index().rename(columns={'index': 'date'})
        df_cat['date'] = df_cat['date'].apply(lambda x: x.strftime('%Y%m%d'))

        if write_excel:
            book = load_workbook(path)
            writer = pd.ExcelWriter(path, engine = 'openpyxl')
            writer.book = book
            df_cat.to_excel(writer, sheet_name = k, index=False)
            book.save(path)
            book.close()
            print("Category {} save in {}.".format(k, path))

        total_df.append(df_cat.set_index('date'))

    return pd.concat(total_df, axis=1, keys=series_dict.keys())


if __name__ == '__main__':
    # # Load credentials
    # credentials = load_credentials("blob_storage")
    # # Start client
    # blob_service_client = BlobServiceClient.from_connection_string(credentials['conn_string'])

    # Get unemployment rate
    series_dict = {
        'Nation Income & Expenditures': ['GDPC1', 'GDPPOT', 'W875RX1', 'PCEC96', 'PSAVERT', 'FYFR', 'FYONET', 'FYFSD',
                                         'GFDEBTN'],
        'Pop Employment Labor': ['PAYEMS', 'UNRATE', 'ICSA', 'UEMPMEAN', 'JTSJOL', 'AWHMAN', 'AHETPI', 'OPHNFB', 'POP',
                                 'CLF16OV', 'CIVPART'],
        'Prod & Bus Act': ['INDPRO', 'TCU', 'BUSINV', 'RRSFS', 'ALTSALES', 'DGORDER', 'BUSLOANS', 'TOTALSL', 'CP', 'HOUST',
                           'PERMIT','UNDCONTSA'],
        'Prices': ['CPIAUCSL', 'PCEPI', 'PCEPILFE', 'GDPDEF', 'PPIFIS', 'WPSFD49207', 'WPSFD4131', 'WPSID62', 'USSTHPI',
                   'SPCS20RSA', 'DCOILWTICO', 'GASREGW', 'MHHNGSP'],
        'Money Bank Finance': ['BOGMBASE', 'M1SL', 'M2SL', 'SP500', 'DJIA', 'WILL5000IND', 'VIXCLS', 'STLFSI2',
                               'BAMLCC0A2AATRIV', 'FF', 'WGS3MO', 'WGS1YR', 'WGS5YR', 'WFII5', 'WGS10YR', 'WFII10', 'WAAA',
                               'WBAA', 'MORTGAGE15US', 'MORTGAGE30US', 'DEXUSEU', 'DEXCHUS', 'DEXCAUS']
    }
    seriesid = ["LNS14000000", "CEU0000000001", "CES0000000001"]
    #df = get_external_data(seriesid)
    df = get_fred_data(series_dict).droplevel(0, axis=1).reset_index()

    # Upload data

    #blob_client = upload_df_parquet(df, "external_data_fred.parquet", blob_service_client, container='external-data')
    # blob_client = upload_df_csv(df,
    #                             "external_data_fred.csv",
    #                             blob_service_client,
    #                             container='external-data')
    # Save clean data in local path
    external_path = "./data/external/"
    if not os.path.exists(external_path):
        os.makedirs(external_path)
    df.to_csv(external_path+"external_data_fred.csv", index=False)
