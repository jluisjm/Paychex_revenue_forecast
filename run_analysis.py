import os, pickle, yaml
import plotly.express as px
from pycaret.regression import *
from datetime import datetime
import seaborn as sns
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook
import src.paychex_ml.data_loader as dl
import src.paychex_ml.models as models

with open("line_items.yml", "r") as stream:
    items_dicctionary = yaml.safe_load(stream)

with open("parameters.yml", "r") as stream:
    locals().update(yaml.safe_load(stream))

ml_col = 'ML Predicted'
uts_col = 'UTS Predicted'
plan_col = 'Plan'
fcst_col = 'Forecast'
fcst_cols = [plan_col, fcst_col, ml_col, uts_col]

file_path = "./data/clean/table_predictable.csv"
drive_path = "./data/clean/table_drivers.csv"
external_path = "./data/external/external_data_fred.csv"

# Set manually date if is necessary
model_run_date = datetime.today().strftime('%Y%m%d')


def print_menu():
    print("000", "--", "All items")
    for key in items_dicctionary.keys():
        print(key, '--', items_dicctionary[key][0])


if __name__ == "__main__":

    print_menu()
    target_col_id = input("Select the target column id: ")

    if target_col_id == '000':
        it = items_dicctionary.values()
        print("Running model training all items")
    else:
        it = [items_dicctionary[target_col_id]]
        print("Running model for {}".format(items_dicctionary[target_col_id]))

    predictions_path = "./data/predictions/" + model_run_date
    if not os.path.exists(predictions_path):
        os.makedirs(predictions_path)
        print("Directory created")

    figures_path = "./data/figures/" + model_run_date
    if not os.path.exists(figures_path):
        os.makedirs(figures_path)
        print("Directory created")

    metadata_path = "./data/metadata/"
    if not os.path.exists(metadata_path):
        os.makedirs(metadata_path)
        print("Directory created")

    model_path = "./data/models/" + model_run_date
    if not os.path.exists(model_path):
        os.makedirs(model_path)
        print("Directory created")

    for target_col, has_drivers, level in it:

        all_df = dl.get_clean_data(train_start_dt, pred_end_dt, file_path, level=level)
        all_df = all_df[['Calendar Date', target_col]]
        if has_drivers:
            driv_df = dl.get_clean_driver_data(train_start_dt, pred_end_dt, target_col, drive_path)
            all_df = pd.merge(all_df, driv_df, on='Calendar Date', how='inner')

        ext_df = pd.read_csv(external_path, dtype={'date': str}) \
            .rename(columns={'date': 'Calendar Date'})

        all_df = pd.merge(all_df, ext_df, on='Calendar Date', how='inner')

        train_df, test_df, comb_df = models.train_test_combine_split(all_df, train_end_dt, test_start_dt, test_end_dt)

        feature_cols = comb_df.columns.to_list()
        feature_cols.remove('Calendar Date')
        feature_cols.remove(target_col)

        # Run Correlations to target
        corr_feature_cols, fig = models.features_correlation(comb_df, target_col, correlation_threshold)
        fig.savefig(figures_path + "/{}_correlations.png".format(target_col))

        # run auto ml and get the most important features
        best = models.run_auto_ml(train_df, test_df, target_col, feature_cols, False, ml_criteria)

        # Here we need to figure out which set of features we want to use
        # ml_features or corr_features
        if feature_selection == 'ml_features':
            ml_features = models.get_important_features('xgboost', features_threshold)
            feature_cols = ml_features['Feature'].tolist()
        elif feature_selection == 'corr_features':
            feature_cols = corr_feature_cols
        else:
            print("No feature selection. This could take a while!")

        if len(ml_features.index) != 0:
            # plot the most important features
            fig = px.bar(ml_features.sort_values('Variable Importance', ascending=True),
                         x='Variable Importance',
                         y='Feature',
                         orientation='h',
                         title='Feature Importance Plot',
                         width=800, height=400)
            fig.write_image(figures_path + "/{}_featureimportance.png".format(target_col))

        with open(model_path + '/{}_features.pkl'.format(target_col), "wb") as fp:  # Pickling
            pickle.dump(feature_cols, fp)
        # ------------------------------------------------------------------------------------- #
        # re-run the auto ml with only the important features
        # ------------------------------------------------------------------------------------- #
        # keeps = ['Calendar Date', target_col]+feature_cols
        keeps = [target_col] + feature_cols
        train_df = train_df[keeps]
        test_df = test_df[keeps]
        # comb_df = comb_df[keeps]

        best = models.run_auto_ml(train_df, test_df, target_col, feature_cols, False, ml_criteria)
        dt_results = pull()

        scores_path = metadata_path + model_run_date + "_scores.xlsx"
        if os.path.exists(scores_path):
            book = load_workbook(scores_path)
        else:
            book = Workbook()
        with pd.ExcelWriter(scores_path, engine='openpyxl') as writer:
            writer.book = book
            dt_results.to_excel(writer, sheet_name=target_col[0:30], index=False)

        # ------------------------------------------------------------------------------------- #
        # generate and plot predicted values on the original dataset
        # ------------------------------------------------------------------------------------- #

        predictions = predict_model(best, data=comb_df)
        predictions['Date'] = pd.date_range(start=str(train_start_dt), end=str(test_end_dt), freq='MS')
        predictions.rename(columns={'Label': target_col + ' - ML Predicted'}, inplace=True)
        fig = px.line(predictions, x='Date', y=[target_col, target_col + ' - ML Predicted'],
                      template='plotly_white',
                      width=800, height=400)
        fig.write_image(figures_path + "/{}_prediction.png".format(target_col))

        final_best = finalize_model(best)
        pipeline, name = save_model(final_best, model_path + '/{}_model'.format(target_col))

        # create the future predictions dataframe
        if has_actuals:
            act_df = all_df[all_df['Calendar Date'] >= pred_start_dt]
            act_df = act_df[['Calendar Date', target_col]]
            act_df['Calendar Date'] = pd.to_datetime(act_df['Calendar Date'])
            pred_df, _ = models.run_auto_arima(comb_df, feature_cols, pred_start_dt, forecast_window, ci=False)
            pred_df = predict_model(final_best, data=pred_df)
            pred_df = pred_df.rename(columns={'Label': ml_col})[['Calendar Date', ml_col]]
            concat_df = pd.merge(act_df, pred_df, on='Calendar Date', how='inner')

            # get plan data
            plan_df = dl.get_clean_data(train_start_dt, pred_end_dt, file_path, type='plan', level=level)
            plan_df = plan_df[['Calendar Date', target_col]]
            plan_df.rename(columns={target_col: plan_col}, inplace=True)
            plan_df['Calendar Date'] = pd.to_datetime(plan_df['Calendar Date'])
            concat_df = pd.merge(concat_df, plan_df, on='Calendar Date', how='inner')

            # get forecast data
            fcst_df = dl.get_clean_data(train_start_dt, pred_end_dt, file_path,
                                        type='forecast',
                                        forecast_type=forecast_type,
                                        level=level)
            fcst_df = fcst_df[['Calendar Date', target_col]]
            fcst_df.rename(columns={target_col: fcst_col}, inplace=True)
            fcst_df['Calendar Date'] = pd.to_datetime(fcst_df['Calendar Date'])
            concat_df = pd.merge(concat_df, fcst_df, on='Calendar Date', how='inner')

        # run UTS
        uts_df = comb_df[['Calendar Date', target_col]]
        uts_df, uts_model = models.run_auto_arima(uts_df, [target_col], pred_start_dt, forecast_window, ci=True,
                                                  alpha=alpha)
        uts_df.rename(columns={target_col: uts_col}, inplace=True)
        concat_df = pd.merge(concat_df, uts_df, on='Calendar Date', how='inner')

        with open(model_path + '/{}_uts_model.pkl'.format(target_col), 'wb') as pkl:
            pickle.dump(uts_model, pkl)

        # combine all data together
        concat_df = pd.concat([comb_df[['Calendar Date', target_col]], concat_df], axis=0)

        # compute mape_df
        mape_df = concat_df[['Calendar Date', target_col, ml_col, uts_col, plan_col, fcst_col]]
        mape_df = mape_df[mape_df['Calendar Date'] >= datetime.strptime(pred_start_dt, '%Y%m%d')]
        mape_df = models.compute_apes_and_mapes(mape_df, 'Calendar Date', target_col, fcst_cols)
        mape_df = mape_df.rename(index={True: 'MAPE'})
        mape_df = pd.concat([mape_df.reset_index(drop=True), uts_df], axis=1)

        mape_path = metadata_path + model_run_date + "_mape.xlsx"
        if os.path.exists(mape_path):
            book = load_workbook(mape_path)
        else:
            book = Workbook()
        with pd.ExcelWriter(mape_path, engine='openpyxl') as writer:
            writer.book = book
            mape_df.to_excel(writer, sheet_name=target_col[0:30], index=False)

        concat_df['Item'] = target_col
        concat_df = concat_df.rename(columns={target_col: 'Actual'}).reset_index(drop=True)

        # show plot
        fig = px.line(concat_df, x='Calendar Date', y=['Actual', ml_col, uts_col, plan_col, fcst_col],
                      template='plotly_white',
                      width=800, height=400)
        fig.write_image(figures_path + "/{}_prediction_impfeat.png".format(target_col))

        concat_df = concat_df[['Calendar Date', 'Item', 'Actual', 'ML Predicted', 'UTS Predicted', 'Plan', 'Forecast',
                               'Lower CI - {}%'.format(100 - alpha * 100),
                               'Upper CI - {}%'.format(100 - alpha * 100)]]
        concat_df = concat_df[concat_df['Calendar Date'] >= pred_start_dt]
        concat_df.to_parquet(predictions_path + "/" + target_col.replace(" ", "") + ".parquet", index=False)
